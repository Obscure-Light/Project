
from __future__ import annotations

import pandas as pd

def export_excel(df, out_path: str):
    # Conditional formatting with openpyxl
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Findings")
        ws = writer.book["Findings"]
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        # Assume Severity in column F (6)
        ws.conditional_formatting.add(f"F2:F{ws.max_row}",
            CellIsRule(operator="equal", formula=['"CRITICAL"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        ws.conditional_formatting.add(f"F2:F{ws.max_row}",
            CellIsRule(operator="equal", formula=['"WARN"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
        )
        ws.conditional_formatting.add(f"F2:F{ws.max_row}",
            CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )
    return out_path
