import socket
import ipaddress
from openpyxl import Workbook

def resolve_addresses(input_lines):
    """
    Accetta una lista di stringhe (ognuna contenente un IP o un hostname).
    Restituisce una lista di tuple (riga_input, ip_risolto, hostname).

    :param input_lines: Lista di stringhe da risolvere (es. IP o hostname)
    :return: Lista di tuple (originale, ip, hostname)
    """
    results = []

    for line in input_lines:
        line = line.strip()
        if not line:
            continue

        ip_address = ""
        hostname = ""
        try:
            # Ricava l'IP da un hostname (se line è un hostname)
            # Se line fosse già un IP, il metodo restituirà l'IP invariato
            ip_address = socket.gethostbyname(line)
        except socket.error:
            # Se l'input sembra un indirizzo IP ma non è risolvibile,
            # manteniamo il valore originale per mostrarlo nel risultato
            try:
                ipaddress.ip_address(line)
                ip_address = line
            except ValueError:
                ip_address = ""
            hostname = "Campo non trovato"
        else:
            # Ricava l'hostname a partire dall'IP ottenuto
            try:
                hostname = socket.gethostbyaddr(ip_address)[0]
            except socket.error:
                hostname = "Campo non trovato"

        results.append((line, ip_address, hostname))

    return results

def save_to_excel(results, output_path):
    """
    Salva i risultati della risoluzione IP/hostname in un file Excel.

    :param results: Lista di tuple (originale, ip, hostname)
    :param output_path: Percorso del file Excel di output
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Risultati"

    # Intestazioni
    sheet['A1'] = "Input"
    sheet['B1'] = "IP"
    sheet['C1'] = "Hostname"

    # Riempimento del foglio
    row_index = 2
    for original, ip_addr, host in results:
        sheet.cell(row=row_index, column=1).value = original
        sheet.cell(row=row_index, column=2).value = ip_addr
        sheet.cell(row=row_index, column=3).value = host
        row_index += 1

    # Salvataggio
    workbook.save(output_path)
